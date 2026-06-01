import pandas as pd
from openpyxl import load_workbook
import re
import os
import shutil


def _normalize_new_pole_spec(text: str) -> str:
    """Normalisasi spesifikasi NEW POLE agar match dengan teks di template.

    Menangani variasi umum dari hasil KMZ/agregator:
    - "NEW POLE - 9m 4" -> "9M 4"
    - "NEW POLE 9 M 4"  -> "9M 4"
    - "NEW POLE 7 2.5"  -> "7M 2.5"
    - "NEW POLE 9M"     -> "9M"
    """
    spec = str(text or "").upper()
    spec = spec.replace("NEW POLE", "").strip()
    spec = spec.replace("-", " ")
    # Buang karakter non-alfanumerik yang sering muncul (mis. 5") atau kurung.
    spec = re.sub(r"[^0-9A-Z.\s]", " ", spec)
    spec = re.sub(r"\s+", " ", spec).strip()

    # Pola umum: "9 4", "9M 4", "9 M 4", "9M4", "9 2.5"
    m = re.match(r"^(\d+)\s*M?\s*(\d+(?:\.\d+)?)$", spec)
    if m:
        return f"{m.group(1)}M {m.group(2)}"

    # Pola umum: "9", "9M", "9 M"
    m = re.match(r"^(\d+)\s*M?$", spec)
    if m:
        return f"{m.group(1)}M"

    return spec

def get_line_info(line_name):
    """Mengekstrak Huruf Line dan Nomor FDT (Cluster Only)"""
    match_line = re.search(r"LINE\s+([A-Z])", str(line_name), re.IGNORECASE)
    line_let = match_line.group(1).upper() if match_line else None
    match_fdt = re.search(r"FDT\s*(\d+)", str(line_name), re.IGNORECASE)
    fdt_num = int(match_fdt.group(1)) if match_fdt else None
    
    # Default ke FDT 1 jika ada Line tapi nomor FDT tidak spesifik
    if line_let and fdt_num is None: 
        fdt_num = 1
    return line_let, fdt_num

def find_row_by_logic(ws, material_search, line_letter=None):
    """Mencari baris berdasarkan deskripsi (B) dan Remark (J) untuk Cluster/Feeder"""
    for row in range(1, ws.max_row + 1):
        desc = str(ws.cell(row=row, column=2).value or "").strip()
        remark = str(ws.cell(row=row, column=10).value or "").strip()
        
        # Validasi Remark (Hanya memproses baris yang memiliki instruksi design)
        if "qty based on design" not in remark.lower(): 
            continue
            
        if line_letter:
            if f"Line {line_letter}".upper() in desc.upper() and material_search.upper() in desc.upper():
                return row
        else:
            if material_search.upper() in desc.upper(): 
                return row
    return None

def run_injection(template_path, source_excel, output_path, mode="CLUSTER"):
    """
    Mesin Injeksi BOQ.
    Membaca data dari source_excel (Step 1) dan menyuntikkannya ke template (Step 2).
    """
    # 1. Validasi Keberadaan File
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"❌ Template backend tidak ditemukan: {template_path}")
    if not os.path.exists(source_excel):
        raise FileNotFoundError(f"❌ File sumber konversi tidak ditemukan: {source_excel}")

    # 2. Persiapan File Output (Copy dari Template agar Original di Assets tetap aman)
    shutil.copyfile(template_path, output_path)

    # 3. Load Data Konversi (Pandas)
    df_boq = pd.read_excel(source_excel, sheet_name="BOQ_PER_LINE")
    df_boq['QTY'] = pd.to_numeric(df_boq['QTY'].replace('-', '0'), errors='coerce').fillna(0)
    df_boq['TOTAL_ROUTE'] = pd.to_numeric(df_boq['TOTAL_ROUTE'].replace('-', '0'), errors='coerce').fillna(0)
    
    # 4. Load Workbook Output (Openpyxl)
    wb = load_workbook(output_path)
    ws = wb["BoM AE"]

    current_mode = mode.upper()

    # ==========================================================
    # LOGIKA MODE FEEDER (Lookup Kolom B, Input Kolom C)
    # ==========================================================
    if current_mode == "FEEDER":
        # Hitung total Joint Closure
        jc_total = df_boq[df_boq['MATERIAL'].str.contains("JOINT", case=False, na=False)]['QTY'].sum()

        # Filter Kabel & Slack
        cable_df = df_boq[
            df_boq['MATERIAL'].str.contains(r'\d+C', regex=True, na=False) & 
            ~df_boq['MATERIAL'].str.contains("SLACK", case=False, na=False)
        ]
        slack_df = df_boq[df_boq['MATERIAL'].str.contains("SLACK", case=False, na=False)]

        # --- A. INJEKSI KABEL & JOINT CLOSURE ---
        for _, c_row in cable_df.iterrows():
            mat_name = str(c_row['MATERIAL']).upper()
            route_val = float(c_row['TOTAL_ROUTE'])
            
            core_match = re.search(r"(\d+)", mat_name)
            core_val = core_match.group(1) if core_match else ""
            search_str = f"{core_val}C/"

            # Cable Injeksi (Mapping Baris 2-8)
            for r in range(2, 9):
                cell_b = str(ws.cell(row=r, column=2).value or "").upper()
                if search_str in cell_b:
                    ws.cell(row=r, column=3).value = route_val
                    break 
            
            # Joint Closure Injeksi (Mapping Baris 23-29)
            for r in range(23, 30):
                cell_b = str(ws.cell(row=r, column=2).value or "").upper()
                if core_val in cell_b and "CORE" in cell_b:
                    ws.cell(row=r, column=3).value = int(jc_total)
                    break 

        # --- B. INJEKSI SLACK (Offset +14 dari baris Kabel) ---
        for _, s_row in slack_df.iterrows():
            slack_mat = str(s_row['MATERIAL']).upper()
            slack_qty = int(s_row['QTY'])
            
            core_match = re.search(r"(\d+)C", slack_mat)
            if core_match:
                search_str = f"{core_match.group(1)}C/"
                for r in range(2, 9):
                    cell_b = str(ws.cell(row=r, column=2).value or "").upper()
                    if search_str in cell_b:
                        ws.cell(row=r + 14, column=3).value = slack_qty
                        break

        # --- C. ENTITAS POLE (Baris 49-59) ---
        for _, p_row in df_boq.iterrows():
            m_name = str(p_row['MATERIAL']).upper()
            if "POLE" in m_name:
                if m_name == "EXISTING POLE":
                    pole_type = "EXISTING POLE EMR"
                elif "NEW POLE" in m_name:
                    pole_type = _normalize_new_pole_spec(m_name)
                else:
                    # Fallback jika ada variasi text lain
                    pole_type = m_name.replace("-", " ")
                    pole_type = re.sub(r"\s+", " ", pole_type).strip()
                
                for r in range(49, 60):
                    cell_b_raw = str(ws.cell(row=r, column=2).value or "")
                    cell_b_upper = cell_b_raw.upper()

                    # Untuk NEW POLE, samakan format input vs template.
                    # Contoh template: "NEW POLE 9-4" harus match input yang ternormalisasi jadi "9M 4".
                    if "NEW POLE" in m_name:
                        if _normalize_new_pole_spec(cell_b_raw) != pole_type:
                            continue
                    else:
                        if pole_type not in cell_b_upper:
                            continue

                    curr = ws.cell(row=r, column=3).value or 0
                    ws.cell(row=r, column=3).value = curr + int(p_row['QTY'])
                    break

    # ==========================================================
    # LOGIKA MODE CLUSTER (Dynamic Column Mapping)
    # ==========================================================
    else:
        # Load Summary FDT untuk penentuan kolom
        try:
            df_fdt = pd.read_excel(source_excel, sheet_name="FDT_SUMMARY")
            for i, row_fdt in df_fdt.iterrows():
                fdt_num = i + 1 
                core = str(row_fdt['CORE'])
                target_row = find_row_by_logic(ws, f"FDT {core} Core")
                if target_row: 
                    ws.cell(row=target_row, column=3 + fdt_num).value = 1
        except:
            print("Warning: Sheet FDT_SUMMARY tidak ditemukan atau kosong.")

        # Injeksi Material per Line ke kolom FDT terkait
        for _, row in df_boq.iterrows():
            line_let, fdt_num = get_line_info(row['LINE'])
            if not fdt_num or fdt_num > 10: 
                continue
                
            material = str(row['MATERIAL']).upper()
            qty, route = row['QTY'], row['TOTAL_ROUTE']
            col_main, col_sling = 3 + fdt_num, 11 + fdt_num 

            if "C/" in material: 
                r = find_row_by_logic(ws, material, line_let)
                if r: ws.cell(row=r, column=col_main).value = float(route)
            elif "SLING WIRE" in material:
                r = find_row_by_logic(ws, "C/", line_let)
                if r: ws.cell(row=r, column=col_sling).value = float(route)
            elif "FAT" in material:
                r_fat_line = find_row_by_logic(ws, f"FAT - Line {line_let}")
                if r_fat_line: ws.cell(row=r_fat_line, column=col_main).value = int(qty)
            elif "NEW POLE" in material:
                search_term = _normalize_new_pole_spec(material)
                r = find_row_by_logic(ws, search_term)
                if r:
                    curr = ws.cell(row=r, column=col_main).value or 0
                    ws.cell(row=r, column=col_main).value = curr + int(qty)
            elif "EXISTING POLE" in material:
                r = find_row_by_logic(ws, "Existing Pole MR")
                if r:
                    curr = ws.cell(row=r, column=col_main).value or 0
                    ws.cell(row=r, column=col_main).value = curr + int(qty)

    # 5. Finalize
    wb.save(output_path)