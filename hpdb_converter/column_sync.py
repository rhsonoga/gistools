from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import re
import os


SOURCE_START_ROW = 10
AV_COLUMN = 48
H_COLUMN = 8
I_COLUMN = 9
J_COLUMN = 10
K_COLUMN = 11
B_COLUMN = 2
A_COLUMN = 1
C_COLUMN = 3
D_COLUMN = 4
E_COLUMN = 5
F_COLUMN = 6
MAX_CORES_PER_TUBE = 10
CORES_PER_TUBE_BLOCK = 12
GROUP_PORTS_PER_BLOCK = 8
FDT_RESET_CODE = "A01"
DEFAULT_MIN_ROW = SOURCE_START_ROW
SHEET_CABLE = "CABLE"
SHEET_FAT_POLE = "FAT & POLE"


def _read_sheet_records(workbook_path, sheet_name, required_fields):
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            return []

        header_index = {
            str(value).strip(): idx
            for idx, value in enumerate(headers)
            if value is not None and str(value).strip()
        }
        missing = [field for field in required_fields if field not in header_index]
        if missing:
            raise KeyError(f"Kolom sheet {sheet_name} tidak lengkap: {', '.join(missing)}")

        records = []
        for row in rows:
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue
            records.append({field: row[header_index[field]] for field in required_fields})
        return records
    finally:
        wb.close()


def _find_last_nonempty_row_in_col(worksheet, col_idx, min_row=DEFAULT_MIN_ROW):
    try:
        max_row = worksheet.max_row or min_row
    except Exception:
        max_row = min_row

    for r in range(max_row, min_row - 1, -1):
        value = worksheet.cell(row=r, column=col_idx).value
        if value is None:
            continue
        if str(value).strip() == "":
            continue
        return r
    return min_row

def run_step3_pole_sync(file_path, parsed_data=None, source_file=None):
    """
    Tahap 3: Sinkronisasi Kolom A-K.
    Memproses Kapasitas, Jalur (Line), Tube, Core, Data Tiang (Pole), dan Grup G.
    """
    
    # --- 1. KONFIGURASI WARNA TUBE (ANSI Standard) ---
    tube_styles = {
        1: ("Blue", "0000FF", "FFFFFF"),
        2: ("Orange", "FF8C00", "000000"),
        3: ("Green", "008000", "FFFFFF"),
        4: ("Brown", "A52A2A", "FFFFFF")
    }

    # File hasil Tahap 1 (sekarang bisa berada di folder temp per-sesi)
    if not source_file:
        source_file = "1.Hasil_Convert.xlsx"  # fallback lama
    source_file = os.path.abspath(source_file)
    if not os.path.exists(source_file):
        raise FileNotFoundError(
            f"❌ Sumber data {source_file} tidak ditemukan. Pastikan Tahap 1 sudah selesai."
        )

    # --- 2. PRE-PROCESS DATA DARI SHEET CABLE (CAPACITY & LINE) ---
    cable_list = []
    line_list = []
    try:
        cable_rows = _read_sheet_records(source_file, SHEET_CABLE, ['Name'])
        for row in cable_rows:
            name_str = str(row['Name'])
            cap_match = re.search(r'(\d+C/\d+T)', name_str)
            line_match = re.search(r'(LINE\s+[A-Z])', name_str.upper())

            cable_list.append(cap_match.group(1) if cap_match else "")
            line_list.append(line_match.group(1) if line_match else "")
            
    except Exception as e:
        print(f"WARNING: Gagal memproses data Capacity/Line: {e}")

    # --- 3. AMBIL DATA POLE DARI HASIL TAHAP 1 ---
    try:
        pole_data = []
        for row in _read_sheet_records(source_file, SHEET_FAT_POLE, ['POLE Name', 'Latitude', 'Longitude']):
            if row.get('POLE Name') is not None and str(row.get('POLE Name')).strip() != "":
                pole_data.append({
                    'name': row['POLE Name'],
                    'lat': row['Latitude'],
                    'long': row['Longitude']
                })
    except Exception as e:
        raise Exception(f"Gagal membaca sumber data tiang: {e}")

    # --- 4. LOAD HPDB UNTUK PROSES INJEKSI FINAL ---
    wb = load_workbook(file_path)
    ws = wb["HOMEPASS DATABASE"]

    # Deteksi baris terakhir yang valid berdasarkan kolom AV (FAT Code)
    # Sebelumnya pakai pandas.read_excel (mahal). Di sini cukup scan dari bawah via openpyxl.
    real_max_row = _find_last_nonempty_row_in_col(ws, col_idx=AV_COLUMN, min_row=SOURCE_START_ROW)

    # Variabel Kontrol Logika
    p_idx = -1
    c_idx = -1
    current_pole = None
    previous_av = None
    current_prefix = None
    fdt_port_counter = 0 
    core_counter = 1
    cores_in_tube = 0
    g_number = 1
    g_occurrence = 0

    # --- 5. LOOPING INJEKSI DATA (BARIS 10 KE BAWAH) ---
    for r in range(SOURCE_START_ROW, real_max_row + 1):
        av_value = ws.cell(row=r, column=AV_COLUMN).value  # Kolom AV
        h_value = ws.cell(row=r, column=H_COLUMN).value    # Kolom H

        if av_value is None or str(av_value).strip() == "":
            continue

        # A. LOGIKA PREFIX & SYNC TIANG
        prefix = str(av_value)[0].upper()
        
        # Reset counter port jika terdeteksi FAT awal (A01)
        if str(av_value).strip().upper() == "A01" and previous_av != "A01":
            fdt_port_counter = 0

        # Reset Core jika Prefix kabel berganti (A, B, C...)
        if prefix != current_prefix:
            core_counter = 1
            cores_in_tube = 0
            current_prefix = prefix
            c_idx += 1 

        # Update data tiang jika kode FAT berganti
        if av_value != previous_av:
            p_idx += 1
            previous_av = av_value
            current_pole = pole_data[p_idx] if p_idx < len(pole_data) else None

        # Injeksi Data Tiang ke Kolom I, J, K
        if current_pole:
            ws.cell(row=r, column=I_COLUMN).value = current_pole['name']
            ws.cell(row=r, column=J_COLUMN).value = current_pole['lat']
            ws.cell(row=r, column=K_COLUMN).value = current_pole['long']
        else:
            ws.cell(row=r, column=I_COLUMN).value = "N/A"

        # B. LOGIKA PENOMORAN PORT (B) & GRUP G (A)
        if h_value in (1, 2):
            fdt_port_counter += 1
            ws.cell(row=r, column=B_COLUMN).value = fdt_port_counter

            if fdt_port_counter == 1:
                g_number = 1
                g_occurrence = 0
            
            # Isi Kolom A (Grup G) hanya pada baris ganjil
            if fdt_port_counter % 2 != 0:
                ws.cell(row=r, column=A_COLUMN).value = f"G{g_number}"
                g_occurrence += 1
                
                # Ganti grup setiap 16 port (8 ganjil)
                if g_occurrence >= GROUP_PORTS_PER_BLOCK:
                    g_number += 1
                    g_occurrence = 0
            else:
                ws.cell(row=r, column=A_COLUMN).value = None

            # C. LOGIKA TUBE & CORE
            tube_num = ((core_counter - 1) // CORES_PER_TUBE_BLOCK) + 1
            
            # Injeksi Capacity & Line pada port pertama tiap kabel
            if h_value == 1:
                if 0 <= c_idx < len(cable_list):
                    ws.cell(row=r, column=C_COLUMN).value = line_list[c_idx]
                    ws.cell(row=r, column=D_COLUMN).value = cable_list[c_idx]

            ws.cell(row=r, column=F_COLUMN).value = core_counter
            
            # Visual Styling Tube (Kolom E)
            if tube_num in tube_styles:
                _, bg_hex, font_hex = tube_styles[tube_num]
                cell_e = ws.cell(row=r, column=E_COLUMN)
                cell_e.value = tube_num 
                cell_e.fill = PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type="solid")
                cell_e.font = Font(color=font_hex, bold=True)

            cores_in_tube += 1
            core_counter += 1

            # Logika pembatasan core per tube (Max 10, skip 2 core cadangan)
            if cores_in_tube == MAX_CORES_PER_TUBE:
                core_counter += 2 
                cores_in_tube = 0

    # --- 6. SIMPAN HASIL FINAL ---
    wb.save(file_path)
    return file_path
