import os
import shutil
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule

MASTER_COLUMNS = ('S', 'T', 'U', 'V', 'W', 'AL', 'AR', 'M', 'AD')
ROW10_START_COL = 53
ROW10_END_COL = 63

def run_step2_basic(parsed_data, template_path, output_path=None):
    """
    REVISI: Menerima template_path (.bin) dari backend, menyalinnya ke 
    .xlsx di root agar bisa diproses oleh openpyxl.
    """
    # 1. Tentukan Nama File Output
    if not output_path:
        # Backward-compatible fallback (akan menulis ke CWD).
        output_path = "2.HPDB_FINAL.xlsx"
    output_name = os.path.abspath(output_path)
    os.makedirs(os.path.dirname(output_name), exist_ok=True)

    # 2. Validasi & Copy dari Backend ke Root
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"❌ File template backend tidak ditemukan di: {template_path}")
    
    # Salin file .bin menjadi .xlsx agar openpyxl tidak error format
    shutil.copyfile(template_path, output_name)

    # 3. Load Workbook dari file hasil copy
    wb_h = load_workbook(output_name)
    ws_h = wb_h["HOMEPASS DATABASE"]
    
    # 2. Persiapkan Data
    # Mengambil data HP COVER hasil parsing KMZ dari dictionary parsed_data
    if "HP COVER" not in parsed_data:
        raise Exception("Data 'HP COVER' tidak ditemukan dalam hasil konversi KMZ.")
        
    hp_list = parsed_data["HP COVER"].to_dict('records')
    
    # 3. Kunci Master Row 10
    # Mengambil nilai default dari baris 10 untuk diduplikasi ke baris-baris berikutnya
    masters = {col: ws_h[f'{col}10'].value for col in MASTER_COLUMNS}
    # Master data untuk range kolom BA sampai BJ (kolom 53-62)
    ba_bj = [ws_h.cell(row=10, column=c).value for c in range(ROW10_START_COL, ROW10_END_COL)]

    # Variabel pembantu untuk logika Counter Kolom H (Nomor Urut Per FAT)
    last_g_id = None
    counter_h = 1

    # 4. Looping Injeksi Data
    for i, item in enumerate(hp_list):
        row = 10 + i
        
        # A. Injeksi Koordinat & Identitas dari KMZ
        ws_h[f'AX{row}'] = item.get('Latitude', '')
        ws_h[f'AY{row}'] = item.get('Longitude', '')
        ws_h[f'AV{row}'] = item.get('Folder Name', '') # fatcode
        ws_h[f'AK{row}'] = item.get('Name', '')        # home_number

        # B. Injeksi Master Data (Duplikasi dari baris 10)
        for col in MASTER_COLUMNS: 
            ws_h[f'{col}{row}'] = masters[col]
            
        for idx, c in enumerate(range(ROW10_START_COL, ROW10_END_COL)): 
            ws_h.cell(row=row, column=c).value = ba_bj[idx]
            
        # C. Injeksi Rumus Excel
        ws_h[f'L{row}'] = f'=AD{row} & " " & AE{row}'
        ws_h[f'G{row}'] = f'=IF(AV{row}="", AU{row}, AU{row} & "." & AV{row})'
        
        # --- D. LOGIKA KOLOM H (Nomor Urut Angka Murni) ---
        # Simulasi nilai G_ID di Python untuk menentukan kapan nomor urut harus reset
        val_au = ws_h[f'AU{row}'].value if ws_h[f'AU{row}'].value else ""
        val_av = item.get('Folder Name', '')
        current_g_id = f"{val_au}.{val_av}" if val_av else f"{val_au}"

        if i == 0:
            counter_h = 1
        else:
            # Jika FAT Code baris ini sama dengan sebelumnya, lanjut hitung
            if current_g_id == last_g_id:
                counter_h += 1
            else:
                # Jika ganti FAT Code, reset urutan ke 1
                counter_h = 1
        
        # Injeksi sebagai Integer agar valid saat dibaca sistem database lain
        ws_h[f'H{row}'] = int(counter_h)
        last_g_id = current_g_id

    # 5. Simpan Hasil Akhir
    wb_h.save(output_name)
    
    return output_name
